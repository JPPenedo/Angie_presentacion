from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def apply_header_style(ws, row, max_col, fill, font):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def main():
    src = r"C:\Users\juanp\Downloads\Vasicek_TreasuryBonds_DTB3.xlsx"
    out = r"C:\Users\juanp\Downloads\Vasicek_TreasuryBonds_DTB3_Entregable.xlsx"
    wb = load_workbook(src)

    orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    soft_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)
    title_font = Font(size=20, bold=True, color="ED7D31")
    subtitle_font = Font(size=12, bold=True, color="7F6000")
    label_font = Font(size=11, bold=True, color="7F6000")
    body_font = Font(size=11, color="404040")

    # Insert cover sheet
    ws_cover = wb.create_sheet("Portada", 0)
    ws_cover.merge_cells("A1:H1")
    ws_cover["A1"] = "Modelo Vasicek - Treasury Bills (DTB3)"
    ws_cover["A1"].font = title_font
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 36

    ws_cover.merge_cells("A2:H2")
    ws_cover["A2"] = "Calibracion con datos hasta diciembre 2025 y verificacion en 2026"
    ws_cover["A2"].font = subtitle_font
    ws_cover["A2"].alignment = Alignment(horizontal="center")
    ws_cover.row_dimensions[2].height = 26

    ws_cover.merge_cells("A4:H4")
    ws_cover["A4"] = "Resumen ejecutivo"
    ws_cover["A4"].fill = orange
    ws_cover["A4"].font = white_bold
    ws_cover["A4"].alignment = Alignment(horizontal="center")

    ws_cover.merge_cells("A5:H5")
    ws_cover["A5"] = (
        "Serie usada: DTB3 (FRED). Modelo: Vasicek de una tasa corta con reversion a la media."
    )
    ws_cover["A5"].font = body_font

    ws_cover.merge_cells("A6:H6")
    ws_cover["A6"] = (
        "El archivo incluye calibracion OLS, escenarios estocasticos para 2026, "
        "comparacion grafica y metricas de error."
    )
    ws_cover["A6"].font = body_font

    ws_cover.merge_cells("A8:H8")
    ws_cover["A8"] = "Ruta de trabajo sugerida"
    ws_cover["A8"].fill = orange
    ws_cover["A8"].font = white_bold
    ws_cover["A8"].alignment = Alignment(horizontal="center")

    steps = [
        "1) Revisar datos historicos en Data_DTB3.",
        "2) Validar parametros en Calibration.",
        "3) Ver escenarios y grafica en Scenarios_2026.",
        "4) Revisar RMSE, MAE y cobertura en Verification.",
        "5) Leer conclusion final en Explicacion.",
    ]
    for i, text in enumerate(steps, start=9):
        ws_cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws_cover[f"A{i}"] = text
        ws_cover[f"A{i}"].font = body_font
        ws_cover[f"A{i}"].alignment = Alignment(horizontal="left")

    for col in "ABCDEFGH":
        ws_cover.column_dimensions[col].width = 18

    # Data sheet style
    ws_data = wb["Data_DTB3"]
    apply_header_style(ws_data, 1, 3, orange, white_bold)
    ws_data.freeze_panes = "A2"
    last_data_row = ws_data.max_row
    add_borders(ws_data, 1, min(last_data_row, 350), 1, 3)

    # Calibration sheet style + legible formula zone
    ws_cal = wb["Calibration"]
    ws_cal["A1"].font = Font(size=16, bold=True, color="ED7D31")

    for r in (3, 7, 12, 17):
        for c in ("A", "B"):
            ws_cal[f"{c}{r}"].fill = orange
            ws_cal[f"{c}{r}"].font = white_bold
            ws_cal[f"{c}{r}"].alignment = Alignment(horizontal="center")

    for r in range(8, 16):
        ws_cal[f"A{r}"].fill = soft_orange
        ws_cal[f"A{r}"].font = label_font
        ws_cal[f"B{r}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws_cal[f"B{r}"].font = Font(name="Consolas", size=10, color="1F1F1F")

    # More legible formula map
    ws_cal.merge_cells("A21:D21")
    ws_cal["A21"] = "Mapa de formulas (formato legible)"
    ws_cal["A21"].fill = orange
    ws_cal["A21"].font = white_bold
    ws_cal["A21"].alignment = Alignment(horizontal="center")

    ws_cal["A22"] = "Parametro"
    ws_cal["B22"] = "Formula matematica"
    ws_cal["C22"] = "Formula en Excel (celda)"
    ws_cal["D22"] = "Descripcion"
    for cell in ("A22", "B22", "C22", "D22"):
        ws_cal[cell].fill = soft_orange
        ws_cal[cell].font = Font(bold=True, color="7F6000")
        ws_cal[cell].alignment = Alignment(horizontal="center", wrap_text=True)

    formula_rows = [
        ("b", "cov(r_t, r_t+1) / var(r_t)", "=SLOPE(y, x)  [B8]", "Persistencia diaria"),
        ("a", "E[r_t+1] - b * E[r_t)", "=INTERCEPT(y, x)  [B9]", "Intercepto AR(1)"),
        ("sigma_e", "desv. estandar residual", "=STEYX(y, x)  [B10]", "Ruido discreto"),
        ("kappa", "-ln(b) / dt", "=-(LN(B8))/B4  [B13]", "Velocidad de reversion"),
        ("theta", "a / (1-b)", "=B9/(1-B8)  [B14]", "Media de largo plazo"),
        ("sigma", "sigma_e * sqrt(2*kappa/(1-b^2))", "=B10*SQRT(2*B13/(1-B8^2))  [B15]", "Volatilidad difusiva"),
    ]

    start = 23
    for i, (p, math_f, excel_f, desc) in enumerate(formula_rows):
        r = start + i
        ws_cal[f"A{r}"] = p
        ws_cal[f"B{r}"] = math_f
        ws_cal[f"C{r}"] = excel_f
        ws_cal[f"D{r}"] = desc
        ws_cal[f"A{r}"].font = label_font
        ws_cal[f"B{r}"].font = Font(name="Cambria", size=11, italic=True)
        ws_cal[f"C{r}"].font = Font(name="Consolas", size=10)
        ws_cal[f"D{r}"].font = body_font
        for c in ("A", "B", "C", "D"):
            ws_cal[f"{c}{r}"].alignment = Alignment(wrap_text=True, vertical="center")
            ws_cal[f"{c}{r}"].fill = PatternFill(
                start_color="FFF8F2", end_color="FFF8F2", fill_type="solid"
            )

    add_borders(ws_cal, 22, 28, 1, 4)
    ws_cal.column_dimensions["A"].width = 34
    ws_cal.column_dimensions["B"].width = 38
    ws_cal.column_dimensions["C"].width = 44
    ws_cal.column_dimensions["D"].width = 30
    for rr in range(23, 29):
        ws_cal.row_dimensions[rr].height = 28

    # Scenarios sheet style
    ws_scen = wb["Scenarios_2026"]
    apply_header_style(ws_scen, 1, 7, orange, white_bold)
    ws_scen.freeze_panes = "A2"
    add_borders(ws_scen, 1, min(ws_scen.max_row, 220), 1, 7)

    # Verification sheet style
    ws_ver = wb["Verification"]
    ws_ver["A1"].font = Font(size=16, bold=True, color="ED7D31")
    apply_header_style(ws_ver, 3, 3, orange, white_bold)
    for r in range(4, 9):
        for c in ("A", "B", "C"):
            ws_ver[f"{c}{r}"].fill = PatternFill(start_color="FFF8F2", end_color="FFF8F2", fill_type="solid")
            ws_ver[f"{c}{r}"].alignment = Alignment(vertical="center", wrap_text=True)
    add_borders(ws_ver, 3, 8, 1, 3)

    # Explanation sheet style
    ws_exp = wb["Explicacion"]
    ws_exp["A1"].font = Font(size=16, bold=True, color="ED7D31")
    for r in (3, 5, 7, 9, 11, 13, 15):
        ws_exp[f"A{r}"].fill = soft_orange
        ws_exp[f"A{r}"].font = label_font
    ws_exp.freeze_panes = "A3"

    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
