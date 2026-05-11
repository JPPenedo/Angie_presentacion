import csv
import datetime as dt
import math
import random
import urllib.request
from statistics import mean

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def percentile(values, q):
    if not values:
        return None
    xs = sorted(values)
    if len(xs) == 1:
        return xs[0]
    pos = (len(xs) - 1) * q
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return xs[lo]
    return xs[lo] + (xs[hi] - xs[lo]) * (pos - lo)


def fetch_dtb3():
    url = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=DTB3"
    with urllib.request.urlopen(url, timeout=30) as resp:
        raw = resp.read().decode("utf-8")

    rows = []
    reader = csv.DictReader(raw.splitlines())
    date_col = reader.fieldnames[0]
    value_col = reader.fieldnames[1]
    for rec in reader:
        d = dt.datetime.strptime(rec[date_col], "%Y-%m-%d").date()
        v = rec[value_col].strip()
        if v == "." or v == "":
            continue
        rate = float(v) / 100.0
        rows.append((d, rate))
    rows.sort(key=lambda x: x[0])
    return rows


def calibrate_vasicek(train_rates, dt_year):
    x = train_rates[:-1]
    y = train_rates[1:]
    mx = mean(x)
    my = mean(y)
    cov_xy = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y))
    var_x = sum((xi - mx) ** 2 for xi in x)
    b = cov_xy / var_x
    a = my - b * mx
    residuals = [yi - (a + b * xi) for xi, yi in zip(x, y)]
    n = len(x)
    sigma_eps = math.sqrt(sum(r * r for r in residuals) / (n - 2))
    kappa = -math.log(b) / dt_year
    theta = a / (1.0 - b)
    sigma = sigma_eps * math.sqrt(2.0 * kappa / (1.0 - b * b))
    return {
        "a": a,
        "b": b,
        "kappa": kappa,
        "theta": theta,
        "sigma_eps": sigma_eps,
        "sigma": sigma,
    }


def simulate_paths(r0, kappa, theta, sigma, dt_year, steps, n_scenarios, seed=2026):
    random.seed(seed)
    exp_k = math.exp(-kappa * dt_year)
    vol_step = sigma * math.sqrt((1.0 - math.exp(-2.0 * kappa * dt_year)) / (2.0 * kappa))
    paths = []
    for _ in range(n_scenarios):
        path = []
        r = r0
        for _ in range(steps):
            z = random.gauss(0.0, 1.0)
            r = theta + (r - theta) * exp_k + vol_step * z
            path.append(r)
        paths.append(path)
    return paths


def main():
    all_rows = fetch_dtb3()
    train_end = dt.date(2025, 12, 31)
    test_start = dt.date(2026, 1, 1)

    train = [r for r in all_rows if r[0] <= train_end]
    test = [r for r in all_rows if r[0] >= test_start]
    if len(train) < 100 or len(test) < 10:
        raise RuntimeError("No hay suficientes datos para calibrar o verificar 2026.")

    dt_year = 1.0 / 252.0
    train_rates = [x[1] for x in train]
    params = calibrate_vasicek(train_rates, dt_year)

    n_scenarios = 500
    steps = len(test)
    r0 = train_rates[-1]
    paths = simulate_paths(
        r0=r0,
        kappa=params["kappa"],
        theta=params["theta"],
        sigma=params["sigma"],
        dt_year=dt_year,
        steps=steps,
        n_scenarios=n_scenarios,
    )

    observed = [x[1] for x in test]
    mean_path, p5, p50, p95, in_band = [], [], [], [], []
    for t in range(steps):
        col = [paths[s][t] for s in range(n_scenarios)]
        m = mean(col)
        q5 = percentile(col, 0.05)
        q50 = percentile(col, 0.50)
        q95 = percentile(col, 0.95)
        mean_path.append(m)
        p5.append(q5)
        p50.append(q50)
        p95.append(q95)
        in_band.append(1 if (observed[t] >= q5 and observed[t] <= q95) else 0)

    rmse = math.sqrt(sum((o - m) ** 2 for o, m in zip(observed, mean_path)) / steps)
    mae = sum(abs(o - m) for o, m in zip(observed, mean_path)) / steps
    coverage = sum(in_band) / steps

    wb = Workbook()

    # Data sheet
    ws_data = wb.active
    ws_data.title = "Data_DTB3"
    ws_data["A1"] = "DATE"
    ws_data["B1"] = "DTB3_rate_decimal"
    ws_data["C1"] = "Sample"
    for c in ("A1", "B1", "C1"):
        ws_data[c].font = Font(bold=True)
        ws_data[c].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    r = 2
    for d, rate in all_rows:
        ws_data.cell(row=r, column=1, value=d)
        ws_data.cell(row=r, column=2, value=rate)
        ws_data.cell(row=r, column=3, value="Train <= 2025-12-31" if d <= train_end else "Test 2026")
        ws_data.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws_data.cell(row=r, column=2).number_format = "0.0000%"
        r += 1
    ws_data.column_dimensions["A"].width = 14
    ws_data.column_dimensions["B"].width = 18
    ws_data.column_dimensions["C"].width = 22

    train_start_row = 2
    train_end_row = 1 + len(train)

    # Calibration sheet
    ws_cal = wb.create_sheet("Calibration")
    ws_cal["A1"] = "Calibracion Vasicek (DTB3)"
    ws_cal["A1"].font = Font(size=14, bold=True)
    ws_cal["A3"] = "Input"
    ws_cal["B3"] = "Value"
    ws_cal["A3"].font = ws_cal["B3"].font = Font(bold=True)
    ws_cal["A4"] = "dt (anual)"
    ws_cal["B4"] = dt_year
    ws_cal["A5"] = "Observaciones train"
    ws_cal["B5"] = len(train) - 1
    ws_cal["B4"].number_format = "0.000000"

    ws_cal["A7"] = "Parametros OLS"
    ws_cal["A7"].font = Font(bold=True)
    ws_cal["A8"] = "b = SLOPE(r_t+1, r_t)"
    ws_cal["B8"] = f"=SLOPE(Data_DTB3!$B$3:$B${train_end_row},Data_DTB3!$B$2:$B${train_end_row-1})"
    ws_cal["A9"] = "a = INTERCEPT(r_t+1, r_t)"
    ws_cal["B9"] = f"=INTERCEPT(Data_DTB3!$B$3:$B${train_end_row},Data_DTB3!$B$2:$B${train_end_row-1})"
    ws_cal["A10"] = "sigma_epsilon = STEYX"
    ws_cal["B10"] = f"=STEYX(Data_DTB3!$B$3:$B${train_end_row},Data_DTB3!$B$2:$B${train_end_row-1})"

    ws_cal["A12"] = "Parametros Vasicek"
    ws_cal["A12"].font = Font(bold=True)
    ws_cal["A13"] = "kappa = -LN(b)/dt"
    ws_cal["B13"] = "=-(LN(B8))/B4"
    ws_cal["A14"] = "theta = a/(1-b)"
    ws_cal["B14"] = "=B9/(1-B8)"
    ws_cal["A15"] = "sigma = sigma_e * SQRT(2*kappa/(1-b^2))"
    ws_cal["B15"] = "=B10*SQRT(2*B13/(1-B8^2))"

    for rr in range(8, 16):
        ws_cal[f"B{rr}"].number_format = "0.000000"

    ws_cal["D3"] = "Estimado (Python)"
    ws_cal["D3"].font = Font(bold=True)
    ws_cal["C8"] = "b"
    ws_cal["D8"] = params["b"]
    ws_cal["C9"] = "a"
    ws_cal["D9"] = params["a"]
    ws_cal["C10"] = "sigma_epsilon"
    ws_cal["D10"] = params["sigma_eps"]
    ws_cal["C13"] = "kappa"
    ws_cal["D13"] = params["kappa"]
    ws_cal["C14"] = "theta"
    ws_cal["D14"] = params["theta"]
    ws_cal["C15"] = "sigma"
    ws_cal["D15"] = params["sigma"]
    for rr in (8, 9, 10, 13, 14, 15):
        ws_cal[f"D{rr}"].number_format = "0.000000"

    ws_cal["A17"] = "Modelo discreto usado para simulacion:"
    ws_cal["A17"].font = Font(bold=True)
    ws_cal["A18"] = "r(t+1)=theta+(r(t)-theta)*exp(-kappa*dt)+sigma*sqrt((1-exp(-2*kappa*dt))/(2*kappa))*Z"
    ws_cal.merge_cells("A18:D18")
    ws_cal["A18"].alignment = Alignment(wrap_text=True)
    ws_cal.column_dimensions["A"].width = 45
    ws_cal.column_dimensions["B"].width = 20
    ws_cal.column_dimensions["C"].width = 18
    ws_cal.column_dimensions["D"].width = 20

    # Scenarios sheet
    ws_scen = wb.create_sheet("Scenarios_2026")
    ws_scen["A1"] = "Date"
    ws_scen["B1"] = "Observed"
    ws_scen["C1"] = "Mean Sim"
    ws_scen["D1"] = "P5"
    ws_scen["E1"] = "P50"
    ws_scen["F1"] = "P95"
    ws_scen["G1"] = "In 90% Band"
    for s in range(1, n_scenarios + 1):
        ws_scen.cell(row=1, column=7 + s, value=f"Scen_{s}")
    for c in range(1, 8):
        cell = ws_scen.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for i, (d, obs) in enumerate(test, start=2):
        idx = i - 2
        ws_scen.cell(row=i, column=1, value=d)
        ws_scen.cell(row=i, column=2, value=obs)
        ws_scen.cell(row=i, column=3, value=mean_path[idx])
        ws_scen.cell(row=i, column=4, value=p5[idx])
        ws_scen.cell(row=i, column=5, value=p50[idx])
        ws_scen.cell(row=i, column=6, value=p95[idx])
        ws_scen.cell(row=i, column=7, value=in_band[idx])
        ws_scen.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        for col in (2, 3, 4, 5, 6):
            ws_scen.cell(row=i, column=col).number_format = "0.0000%"
        for s in range(n_scenarios):
            ws_scen.cell(row=i, column=8 + s, value=paths[s][idx])
            ws_scen.cell(row=i, column=8 + s).number_format = "0.0000%"

    ws_scen.column_dimensions["A"].width = 14
    ws_scen.column_dimensions["B"].width = 12
    ws_scen.column_dimensions["C"].width = 12
    ws_scen.column_dimensions["D"].width = 10
    ws_scen.column_dimensions["E"].width = 10
    ws_scen.column_dimensions["F"].width = 10
    ws_scen.column_dimensions["G"].width = 12

    last_row = 1 + len(test)

    chart = LineChart()
    chart.title = "DTB3 2026: Observada vs Escenarios Vasicek"
    chart.y_axis.title = "Tasa"
    chart.x_axis.title = "Fecha"
    chart.height = 11
    chart.width = 22
    data = Reference(ws_scen, min_col=2, max_col=6, min_row=1, max_row=last_row)
    cats = Reference(ws_scen, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_scen.add_chart(chart, "A3")

    # Verification sheet
    ws_ver = wb.create_sheet("Verification")
    ws_ver["A1"] = "Verificacion del modelo (2026)"
    ws_ver["A1"].font = Font(size=14, bold=True)
    ws_ver["A3"] = "Metrica"
    ws_ver["B3"] = "Formula en Excel"
    ws_ver["C3"] = "Valor"
    for c in ("A3", "B3", "C3"):
        ws_ver[c].font = Font(bold=True)
        ws_ver[c].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws_ver["A4"] = "RMSE (Observed vs Mean Sim)"
    ws_ver["B4"] = f"=SQRT(SUMXMY2(Scenarios_2026!B2:B{last_row},Scenarios_2026!C2:C{last_row})/COUNTA(Scenarios_2026!B2:B{last_row}))"
    ws_ver["C4"] = rmse

    ws_ver["A5"] = "MAE (Observed vs Mean Sim)"
    ws_ver["B5"] = f"=SUMPRODUCT(ABS(Scenarios_2026!B2:B{last_row}-Scenarios_2026!C2:C{last_row}))/COUNTA(Scenarios_2026!B2:B{last_row})"
    ws_ver["C5"] = mae

    ws_ver["A6"] = "Cobertura banda 5%-95%"
    ws_ver["B6"] = f"=AVERAGE(Scenarios_2026!G2:G{last_row})"
    ws_ver["C6"] = coverage

    ws_ver["A7"] = "Numero de escenarios"
    ws_ver["B7"] = "Input de simulacion"
    ws_ver["C7"] = n_scenarios

    ws_ver["A8"] = "Periodo de verificacion"
    ws_ver["B8"] = "Datos observados"
    ws_ver["C8"] = f"{test[0][0]} a {test[-1][0]}"

    ws_ver["C4"].number_format = "0.0000%"
    ws_ver["C5"].number_format = "0.0000%"
    ws_ver["C6"].number_format = "0.00%"
    ws_ver.column_dimensions["A"].width = 32
    ws_ver.column_dimensions["B"].width = 65
    ws_ver.column_dimensions["C"].width = 18

    # Explanation sheet
    ws_exp = wb.create_sheet("Explicacion")
    ws_exp["A1"] = "Resumen de la actividad: Vasicek con Treasury Bills (DTB3)"
    ws_exp["A1"].font = Font(size=14, bold=True)
    ws_exp.merge_cells("A1:E1")

    text_rows = [
        "1) Serie utilizada y por que:",
        (
            "Se utilizo DTB3 (3-Month Treasury Bill Secondary Market Rate) de FRED "
            "(Board of Governors of the Federal Reserve System). Es una tasa de corto plazo "
            "liquida y estandar para calibrar modelos de corto plazo."
        ),
        "2) Estimacion de parametros:",
        (
            "Se aplico una regresion OLS de r(t+1)=a+b*r(t) con datos hasta 2025-12-31. "
            "Luego se mapearon los parametros de Vasicek: kappa=-ln(b)/dt, "
            "theta=a/(1-b), sigma=sigma_e*sqrt(2*kappa/(1-b^2))."
        ),
        "3) Escenarios y supuestos:",
        (
            f"Se simularon {n_scenarios} trayectorias diarias para 2026 usando el esquema "
            "exacto discreto de Vasicek, con dt=1/252 y punto inicial igual a la ultima "
            "tasa observada de 2025."
        ),
        "4) Comparacion contra observado 2026:",
        (
            "Se comparo la serie observada con la media simulada y las bandas percentiles "
            "5%-95% (grafica en Scenarios_2026). Adicionalmente se reportan RMSE, MAE y "
            "cobertura de banda."
        ),
        "5) Desempeno y limitaciones:",
        (
            "El modelo de Vasicek captura reversión a la media y produce escenarios coherentes, "
            "pero puede fallar ante cambios de regimen, choques macro inesperados y no impone "
            "tasas siempre positivas."
        ),
        "Conclusiones (pregunta final):",
        (
            "El modelo calibrado hasta dic-2025 es razonable como aproximacion base para 2026 "
            "si los errores y la cobertura son aceptables; sin embargo, no reproduce perfectamente "
            "episodios de alta volatilidad o cambios estructurales. Por ello funciona como "
            "benchmark, no como pronostico exacto."
        ),
        "Fuente:",
        "https://fred.stlouisfed.org/series/DTB3",
    ]

    row = 3
    for line in text_rows:
        ws_exp[f"A{row}"] = line
        if line.endswith(":") or line.startswith("Conclusiones") or line == "Fuente:":
            ws_exp[f"A{row}"].font = Font(bold=True)
        ws_exp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws_exp[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws_exp.column_dimensions["A"].width = 120
    for rr in range(3, row):
        ws_exp.row_dimensions[rr].height = 40

    out_path = r"C:\Users\juanp\Downloads\Vasicek_TreasuryBonds_DTB3.xlsx"
    wb.save(out_path)
    print(out_path)
    print(f"train_rows={len(train)} test_rows={len(test)}")
    print(
        "params",
        f"a={params['a']:.6f}",
        f"b={params['b']:.6f}",
        f"kappa={params['kappa']:.4f}",
        f"theta={params['theta']:.4%}",
        f"sigma={params['sigma']:.4%}",
    )
    print(
        "metrics",
        f"rmse={rmse:.4%}",
        f"mae={mae:.4%}",
        f"coverage={coverage:.2%}",
    )


if __name__ == "__main__":
    main()
