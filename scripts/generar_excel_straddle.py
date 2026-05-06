"""
Genera Straddle_sintetico.xlsx con el mismo esquema que Ejemplo.xlsx (bloque Datos + tabla de escenarios).
Ejecutar desde la raíz del repo: python scripts/generar_excel_straddle.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Straddle_sintetico.xlsx"

thin = Side(style="thin", color="FF333333")
border_grid = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_header = PatternFill("solid", fgColor="FFE7E6E6")
fill_datos = PatternFill("solid", fgColor="FFF5F5F5")
font_bold = Font(bold=True)


def style_header_row(ws, row, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_bold
        cell.fill = fill_header
        cell.border = border_grid
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def main():
    wb = Workbook()

    # --- Hoja 1: payoff al vencimiento (formato tipo Ejemplo) ---
    ws1 = wb.active
    ws1.title = "Straddle compra"

    ws1["A1"] = "Datos"
    ws1["A1"].font = Font(bold=True, size=12)

    labels = [
        ("A3", "Strike (K)"),
        ("A4", "Prima call (c)"),
        ("A5", "Prima put (p)"),
        ("A6", "No. opciones "),
        ("A7", "Pago total de prima "),
    ]
    for addr, text in labels:
        ws1[addr] = text
        ws1[addr].font = font_bold

    ws1["B3"] = 100
    ws1["B4"] = 5
    ws1["B5"] = 3
    ws1["B6"] = 100
    ws1["B7"] = "=(B4+B5)*B6"

    for r in range(3, 8):
        ws1.cell(row=r, column=1).fill = fill_datos
        ws1.cell(row=r, column=2).fill = fill_datos

    headers = [
        (9, 4, "Precio de S al vencimiento "),
        (9, 5, "Ganancia - Pérdida Call larga"),
        (9, 6, "Ganancia - Pérdida Put larga"),
        (9, 7, "Resultado straddle (comprador)"),
        (9, 8, "Nota"),
    ]
    for row, col, val in headers:
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = font_bold

    style_header_row(ws1, 9, 4, 8)

    # Precios ST de ejemplo (incluye breakevens 92 y 108 con K=100, c+p=8)
    st_values = [80, 85, 90, 92, 95, 100, 105, 108, 110, 115, 120]
    start_row = 10
    for i, st in enumerate(st_values):
        r = start_row + i
        ws1.cell(row=r, column=4, value=st)
        ws1.cell(row=r, column=5, value=f"=(MAX(0,D{r}-$B$3)-$B$4)*$B$6")
        ws1.cell(row=r, column=6, value=f"=(MAX(0,$B$3-D{r})-$B$5)*$B$6")
        ws1.cell(row=r, column=7, value=f"=E{r}+F{r}")
        for c in range(4, 8):
            ws1.cell(row=r, column=c).border = border_grid

    ws1.cell(row=13, column=8, value="Cerca break-even inferior (~92)")
    ws1.cell(row=16, column=8, value="Pérdida máxima si S_T = K")
    ws1.cell(row=18, column=8, value="Cerca break-even superior (~108)")

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["D"].width = 26
    ws1.column_dimensions["E"].width = 26
    ws1.column_dimensions["F"].width = 26
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["H"].width = 34

    # --- Hoja 2: referencia rápida (como segunda hoja del ejemplo) ---
    ws2 = wb.create_sheet("Resumen fórmulas")
    ws2["A1"] = "Long straddle (compra call + compra put, mismo K y T)"
    ws2["A1"].font = Font(bold=True, size=12)

    rows_txt = [
        ("A3", "Prima total pagada (por acción)", "B3", "c + p"),
        ("A4", "Pérdida máxima (por acción, a vencimiento)", "B4", "c + p si S_T = K"),
        ("A5", "Break-even superior", "B5", "K + (c + p)"),
        ("A6", "Break-even inferior", "B6", "K - (c + p)"),
        ("A7", "Payoff call larga al vencimiento", "B7", "max(0, S_T - K) - c"),
        ("A8", "Payoff put larga al vencimiento", "B8", "max(0, K - S_T) - p"),
        ("A9", "Payoff straddle (por acción)", "B9", "suma de las dos patas anteriores"),
        ("A11", "Ejemplo numérico (presentación)", "B11", "K=100, c=5, p=3, n=100 contratos por pata"),
    ]
    for a, la, b, lb in rows_txt:
        ws2[a] = la
        ws2[a].font = font_bold
        ws2[b] = lb

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 52

    wb.save(OUT)
    print("Guardado:", OUT)


if __name__ == "__main__":
    main()
